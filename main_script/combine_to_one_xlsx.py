import os
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook

input_folder = "../main_script/output/"
output_file = "../main_script/combined_file/combined.xlsx"

if not os.path.exists(input_folder):
    raise FileNotFoundError(f"Input folder '{input_folder}' not found!")

input_files = [file for file in os.listdir(input_folder) if file.endswith('.xlsx')]

if not input_files:
    raise FileNotFoundError("No .xlsx files found in the input folder.")

wb = Workbook()
ws = wb.active
ws.title = "CombinedTables"

current_row = 1

for file in input_files:
    file_path = os.path.join(input_folder, file)
    df = pd.read_excel(file_path)

    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), current_row):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    current_row = ws.max_row + 5

output_dir = os.path.dirname(output_file)
if not os.path.exists(output_dir):
    os.makedirs(output_dir)

wb.save(output_file)

print(f"Combined tables with headers saved to {output_file}")